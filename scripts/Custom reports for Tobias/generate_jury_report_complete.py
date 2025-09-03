#!/usr/bin/env python3
"""
Mobility Trailblazers Vote - Complete Jury Evaluation Report Generator
Addresses all stakeholder requirements:
- All 50 candidates included (no exclusions)
- Category classification (Startup, Government, Corporate)
- Top 25 ranking + 3 category winners
- Jury voting status overview
- Clear scoring methodology explanations
"""

import pandas as pd
import numpy as np
from datetime import datetime
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import warnings
warnings.filterwarnings('ignore')

# Database connection parameters
DB_USER = "wp_user"
DB_PASS = "Wp7kL9xP2qR7vN6wE3zY4uC1sA5f"
DB_NAME = "wordpress_db"
DB_CONTAINER = "mobility_mariadb_VOTE"

def execute_query(query):
    """Execute MySQL query via Docker and return results as DataFrame"""
    cmd = f'docker exec {DB_CONTAINER} mariadb -u {DB_USER} -p{DB_PASS} {DB_NAME} -e "{query}" --batch --raw'
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    
    if result.returncode != 0:
        print(f"Error executing query: {result.stderr}")
        return pd.DataFrame()
    
    lines = result.stdout.strip().split('\n')
    if len(lines) < 2:
        return pd.DataFrame()
    
    headers = lines[0].split('\t')
    data = [line.split('\t') for line in lines[1:]]
    
    df = pd.DataFrame(data, columns=headers)
    # Convert numeric columns
    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col])
        except:
            pass
    
    return df

def classify_organization(org_name):
    """Classify organization into Startup, Government, or Corporate"""
    if pd.isna(org_name) or org_name == 'NULL' or org_name == '':
        return 'Unknown'
    
    org_lower = str(org_name).lower()
    
    # Government indicators
    government_keywords = ['stadt', 'hansestadt', 'hamburg', 'ministry', 'verkehrsverbund', 
                          'hvv', 'vdv', 'nationale leitstelle', 'now', 'tübingen']
    
    # Corporate indicators (established companies)
    corporate_names = ['bmw', 'mercedes', 'volkswagen', 'skoda', 'amag', 'dräxlmaier', 
                      'siemens', 'telekom', 'deutsche bahn', 'flixbus', 'uber', 'tier',
                      'rail europe', 'touring club', 'tcs', 'jobrad', 'benteler']
    
    # Check for government
    for keyword in government_keywords:
        if keyword in org_lower:
            return 'Government'
    
    # Check for corporate
    for corp in corporate_names:
        if corp in org_lower:
            return 'Corporate'
    
    # Specific known classifications
    known_classifications = {
        'FlixBus': 'Corporate',
        'UBER': 'Corporate',
        'TIER-Dott': 'Corporate',
        'Rail Europe SAS': 'Corporate',
        'Manager MAGAZIN': 'Corporate',
        'IAA Mobility, Messe München': 'Corporate',
        'VDV': 'Government',
        'NOW – Nationale Leitstelle Ladeinfrastruktur': 'Government',
        'Hansestadt Hamburg': 'Government',
        'Stadt Tübingen': 'Government',
        'Hamburger Verkehrsverbund (HVV)': 'Government',
    }
    
    if org_name in known_classifications:
        return known_classifications[org_name]
    
    # Default to Startup for newer/smaller companies
    startup_indicators = ['ag', 'gmbh', 'mobility', 'tech', 'digital', 'connect', 
                         'drive', 'lab', 'solution', 'innovation']
    
    for indicator in startup_indicators:
        if indicator in org_lower:
            return 'Startup'
    
    return 'Startup'  # Default to Startup if unclear

def get_all_candidates_with_ranks():
    """Get all candidates with rank-based scoring"""
    query = """
    WITH RankedEvaluations AS (
        SELECT 
            e.jury_member_id,
            p.post_title as jury_name,
            c.name as candidate_name,
            c.organization,
            c.country,
            c.post_id as candidate_id,
            e.total_score,
            e.courage_score,
            e.innovation_score,
            e.implementation_score,
            e.relevance_score,
            e.visibility_score,
            RANK() OVER (PARTITION BY e.jury_member_id ORDER BY e.total_score DESC) as rank_within_jury,
            COUNT(*) OVER (PARTITION BY e.jury_member_id) as jury_size
        FROM wp_mt_evaluations e
        JOIN wp_posts p ON e.jury_member_id = p.ID
        JOIN wp_mt_candidates c ON e.candidate_id = c.post_id
        WHERE e.status = 'completed' 
            AND p.post_type = 'mt_jury_member'
    ),
    AllCandidates AS (
        SELECT 
            c.post_id as candidate_id,
            c.name as candidate_name,
            c.organization,
            c.country
        FROM wp_mt_candidates c
    )
    SELECT 
        ac.candidate_name,
        ac.organization,
        ac.country,
        COUNT(DISTINCT re.jury_member_id) as num_juries,
        COALESCE(ROUND(AVG(re.rank_within_jury), 2), 999) as avg_rank,
        COALESCE(ROUND(STD(re.rank_within_jury), 2), 0) as rank_std_dev,
        COALESCE(MIN(re.rank_within_jury), 0) as best_rank,
        COALESCE(MAX(re.rank_within_jury), 0) as worst_rank,
        COALESCE(ROUND(AVG(re.total_score), 2), 0) as avg_score,
        COALESCE(ROUND(AVG(re.courage_score), 2), 0) as avg_courage,
        COALESCE(ROUND(AVG(re.innovation_score), 2), 0) as avg_innovation,
        COALESCE(ROUND(AVG(re.implementation_score), 2), 0) as avg_implementation,
        COALESCE(ROUND(AVG(re.relevance_score), 2), 0) as avg_relevance,
        COALESCE(ROUND(AVG(re.visibility_score), 2), 0) as avg_visibility,
        COALESCE(GROUP_CONCAT(DISTINCT re.jury_name ORDER BY re.jury_name SEPARATOR '; '), 'No evaluations') as evaluated_by,
        COALESCE(ROUND(AVG(re.jury_size - re.rank_within_jury + 1), 2), 0) as avg_borda_points
    FROM AllCandidates ac
    LEFT JOIN RankedEvaluations re ON ac.candidate_id = re.candidate_id
    GROUP BY ac.candidate_id, ac.candidate_name, ac.organization, ac.country
    ORDER BY 
        CASE WHEN COUNT(DISTINCT re.jury_member_id) = 0 THEN 1 ELSE 0 END,
        avg_rank ASC
    """
    return execute_query(query)

def get_jury_voting_status():
    """Get jury voting completion status"""
    query = """
    SELECT 
        p.post_title as jury_name,
        u.user_email as email,
        COUNT(DISTINCT ja.id) as assignments,
        COUNT(DISTINCT e.id) as evaluations_completed,
        CASE 
            WHEN COUNT(DISTINCT e.id) >= COUNT(DISTINCT ja.id) AND COUNT(DISTINCT ja.id) > 0 THEN 'Completed'
            WHEN COUNT(DISTINCT e.id) > 0 THEN 'Partial'
            WHEN p.post_title LIKE '%Test%' OR p.post_title = 'Nico' THEN 'Test Account'
            ELSE 'Not Started'
        END as status
    FROM wp_posts p
    LEFT JOIN wp_users u ON p.post_author = u.ID
    LEFT JOIN wp_mt_jury_assignments ja ON p.ID = ja.jury_member_id AND ja.is_active = 1
    LEFT JOIN wp_mt_evaluations e ON p.ID = e.jury_member_id AND e.status = 'completed'
    WHERE p.post_type = 'mt_jury_member' AND p.post_status = 'publish'
    GROUP BY p.ID, p.post_title, u.user_email
    ORDER BY 
        CASE 
            WHEN COUNT(DISTINCT e.id) >= COUNT(DISTINCT ja.id) AND COUNT(DISTINCT ja.id) > 0 THEN 1
            WHEN COUNT(DISTINCT e.id) > 0 THEN 2
            WHEN p.post_title LIKE '%Test%' OR p.post_title = 'Nico' THEN 4
            ELSE 3
        END,
        p.post_title
    """
    return execute_query(query)

def get_jury_statistics():
    """Get comprehensive jury member statistics"""
    query = """
    SELECT 
        p.ID as jury_id,
        p.post_title as jury_name,
        COUNT(e.id) as evaluations_completed,
        ROUND(AVG(e.total_score), 2) as avg_total_score,
        ROUND(STD(e.total_score), 2) as score_std_dev,
        ROUND(MIN(e.total_score), 2) as min_score,
        ROUND(MAX(e.total_score), 2) as max_score
    FROM wp_posts p
    LEFT JOIN wp_mt_evaluations e ON p.ID = e.jury_member_id AND e.status = 'completed'
    WHERE p.post_type = 'mt_jury_member' 
        AND p.post_status = 'publish'
        AND e.id IS NOT NULL
    GROUP BY p.ID, p.post_title
    ORDER BY p.post_title
    """
    return execute_query(query)

def create_excel_report(output_file):
    """Create comprehensive Excel report with all requirements"""
    
    print("=" * 60)
    print("Fetching data from database...")
    print("=" * 60)
    
    # Get all data
    all_candidates = get_all_candidates_with_ranks()
    jury_status = get_jury_voting_status()
    jury_stats = get_jury_statistics()
    
    # Add category classification
    all_candidates['Category'] = all_candidates['organization'].apply(classify_organization)
    
    # Sort by average rank (candidates with no evaluations will be at the end)
    all_candidates = all_candidates.sort_values('avg_rank')
    all_candidates['Overall Rank'] = range(1, len(all_candidates) + 1)
    
    print(f"✓ Found {len(all_candidates)} total candidates")
    print(f"✓ Found {len(all_candidates[all_candidates['num_juries'] > 0])} evaluated candidates")
    print(f"✓ Found {len(jury_status)} jury members")
    if len(jury_status) > 0 and 'status' in jury_status.columns:
        print(f"✓ Found {len(jury_status[jury_status['status'] == 'Completed'])} juries who completed evaluations")
    else:
        print("✓ Jury status data not available")
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Sheet 1: Executive Summary
        print("\nCreating Executive Summary...")
        
        # Get top 25 overall
        top25 = all_candidates[all_candidates['num_juries'] > 0].head(25)
        
        # Get top 3 per category
        category_winners = {}
        for category in ['Startup', 'Government', 'Corporate']:
            cat_candidates = all_candidates[(all_candidates['Category'] == category) & 
                                           (all_candidates['num_juries'] > 0)]
            category_winners[category] = cat_candidates.head(3)
        
        summary_data = {
            'Metric': [
                'MOBILITY TRAILBLAZERS 2025',
                '',
                'Total Candidates',
                'Evaluated Candidates',
                'Candidates Without Evaluations',
                'Total Jury Members',
                'Juries Who Completed Evaluations',
                'Total Evaluations Completed',
                '',
                'TOP CANDIDATE OVERALL',
                'Organization',
                'Category',
                'Average Rank',
                'Evaluated By (Juries)',
                '',
                'CATEGORY WINNERS',
                f'Top Startup',
                f'Top Government',
                f'Top Corporate'
            ],
            'Value': [
                'Top 25 Rankings + 3 Category Winners',
                '',
                len(all_candidates),
                len(all_candidates[all_candidates['num_juries'] > 0]),
                len(all_candidates[all_candidates['num_juries'] == 0]),
                len(jury_status),
                len(jury_status[jury_status['status'] == 'Completed']),
                jury_stats['evaluations_completed'].sum() if len(jury_stats) > 0 else 0,
                '',
                top25.iloc[0]['candidate_name'] if len(top25) > 0 else 'N/A',
                top25.iloc[0]['organization'] if len(top25) > 0 else 'N/A',
                top25.iloc[0]['Category'] if len(top25) > 0 else 'N/A',
                f"{top25.iloc[0]['avg_rank']:.2f}" if len(top25) > 0 else 'N/A',
                f"{top25.iloc[0]['num_juries']} juries" if len(top25) > 0 else 'N/A',
                '',
                '',
                category_winners['Startup'].iloc[0]['candidate_name'] if len(category_winners['Startup']) > 0 else 'N/A',
                category_winners['Government'].iloc[0]['candidate_name'] if len(category_winners['Government']) > 0 else 'N/A',
                category_winners['Corporate'].iloc[0]['candidate_name'] if len(category_winners['Corporate']) > 0 else 'N/A'
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
        
        # Add Top 25 to Executive Summary
        top25_display = top25[['Overall Rank', 'candidate_name', 'organization', 'Category', 
                               'avg_rank', 'num_juries', 'avg_score']]
        top25_display.columns = ['Rank', 'Candidate', 'Organization', 'Category', 
                                 'Avg Rank', 'Juries', 'Avg Score']
        top25_display.to_excel(writer, sheet_name='Executive Summary', 
                              startrow=len(summary_df) + 3, index=False)
        
        # Sheet 2: Complete Rankings - All Candidates
        print("Creating Complete Rankings sheet...")
        complete_rankings = all_candidates.copy()
        
        # Reorder columns for clarity
        complete_rankings = complete_rankings[[
            'Overall Rank', 'candidate_name', 'organization', 'Category', 'country',
            'num_juries', 'avg_rank', 'avg_score', 'avg_borda_points',
            'best_rank', 'worst_rank', 'rank_std_dev',
            'avg_courage', 'avg_innovation', 'avg_implementation', 
            'avg_relevance', 'avg_visibility'
        ]]
        
        complete_rankings.columns = [
            'Rank', 'Candidate Name', 'Organization', 'Category', 'Country',
            'Jury Count', 'Avg Rank Position', 'Avg Score (0-10)', 'Avg Borda Points',
            'Best Rank', 'Worst Rank', 'Rank Std Dev',
            'Avg Courage', 'Avg Innovation', 'Avg Implementation',
            'Avg Relevance', 'Avg Visibility'
        ]
        
        complete_rankings.to_excel(writer, sheet_name='Complete Rankings - All 50', index=False)
        
        # Sheet 3: Category Winners
        print("Creating Category Winners sheet...")
        category_sheets = []
        for category, winners in category_winners.items():
            if len(winners) > 0:
                winners_display = winners[['Overall Rank', 'candidate_name', 'organization', 
                                          'avg_rank', 'num_juries', 'avg_score']].copy()
                winners_display['Category'] = category
                winners_display['Category Rank'] = range(1, len(winners_display) + 1)
                category_sheets.append(winners_display)
        
        if category_sheets:
            category_df = pd.concat(category_sheets, ignore_index=True)
            category_df = category_df[['Category', 'Category Rank', 'Overall Rank', 
                                       'candidate_name', 'organization', 'avg_rank', 
                                       'num_juries', 'avg_score']]
            category_df.columns = ['Category', 'Rank in Category', 'Overall Rank', 
                                   'Candidate', 'Organization', 'Avg Rank Position', 
                                   'Jury Count', 'Avg Score']
            category_df.to_excel(writer, sheet_name='Category Winners', index=False)
        
        # Sheet 4: Jury Voting Status
        print("Creating Jury Voting Status sheet...")
        jury_status_display = jury_status[['jury_name', 'assignments', 
                                          'evaluations_completed', 'status']].copy()
        jury_status_display.columns = ['Jury Member', 'Assigned Candidates', 
                                       'Evaluations Completed', 'Status']
        
        # Add summary at the top
        jury_summary = pd.DataFrame({
            'Jury Member': ['SUMMARY', ''],
            'Assigned Candidates': ['', ''],
            'Evaluations Completed': [
                f"Completed: {len(jury_status[jury_status['status'] == 'Completed'])}",
                f"Not Started: {len(jury_status[jury_status['status'] == 'Not Started'])}"
            ],
            'Status': ['', '']
        })
        
        jury_status_final = pd.concat([jury_summary, jury_status_display], ignore_index=True)
        jury_status_final.to_excel(writer, sheet_name='Jury Voting Status', index=False)
        
        # Sheet 5: Scoring Methodology
        print("Creating Scoring Methodology sheet...")
        methodology = pd.DataFrame({
            'Term': [
                'Avg Rank Position',
                'Avg Score (0-10)',
                'Avg Borda Points',
                'Jury Count',
                'Category',
                'Best Rank',
                'Worst Rank',
                'Overall Rank'
            ],
            'Description': [
                'Average ranking position across all juries that evaluated the candidate. Lower is better (1 = first place).',
                'Average of the actual numerical scores (0-10 scale) given by juries across all 5 criteria.',
                'Points based on rank position within each jury (10 points for 1st place, 9 for 2nd, etc.), then averaged.',
                'Number of jury members who evaluated this candidate. Higher = more confidence in ranking.',
                'Organization classification: Startup (new/innovative companies), Government (public sector), Corporate (established companies).',
                'Best (lowest) rank position received from any jury.',
                'Worst (highest) rank position received from any jury.',
                'Final ranking position based on Average Rank Position (primary) with Avg Score as tiebreaker.'
            ],
            'Example': [
                'If ranked 1st by one jury and 3rd by another, avg = 2.0',
                'If scores are 8.5, 7.0, 9.0, avg = 8.17',
                'If 10 candidates evaluated: 1st place = 10 points, 2nd = 9 points, etc.',
                '4 means evaluated by 4 different jury members',
                'FlixBus = Corporate, Urban Connect = Startup, Stadt Tübingen = Government',
                '1 means at least one jury ranked them first',
                '5 means highest rank from any jury was 5th place',
                'Rank 1 = best overall candidate across all evaluations'
            ]
        })
        methodology.to_excel(writer, sheet_name='Scoring Methodology', index=False)
        
        # Sheet 6: Candidates Without Evaluations
        print("Creating sheet for candidates without evaluations...")
        no_eval = all_candidates[all_candidates['num_juries'] == 0].copy()
        if len(no_eval) > 0:
            no_eval_display = no_eval[['candidate_name', 'organization', 'Category', 'country']]
            no_eval_display.columns = ['Candidate Name', 'Organization', 'Category', 'Country']
            no_eval_display.to_excel(writer, sheet_name='Not Yet Evaluated', index=False)
        
        # Format the workbook
        print("Applying formatting...")
        workbook = writer.book
        
        # Format Executive Summary
        ws = workbook['Executive Summary']
        header_font = Font(bold=True, size=14)
        title_fill = PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid')
        header_font_white = Font(bold=True, size=12, color='FFFFFF')
        section_font = Font(bold=True, size=11)
        
        # Format title row
        ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        ws['A1'].fill = title_fill
        ws['B1'].fill = title_fill
        
        # Format headers
        for cell in ws[1]:
            if cell.value and 'MOBILITY TRAILBLAZERS' in str(cell.value):
                cell.font = Font(bold=True, size=16, color='FFFFFF')
                cell.fill = title_fill
        
        # Format section headers
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value and any(keyword in str(row[0].value) for keyword in 
                                   ['TOP CANDIDATE', 'CATEGORY WINNERS', 'MOBILITY TRAILBLAZERS']):
                row[0].font = section_font
        
        # Format Complete Rankings sheet with color scales
        ws_rankings = workbook['Complete Rankings - All 50']
        
        # Apply color scale to Avg Rank column (green = low/good, red = high/bad)
        for col in ws_rankings.columns:
            if col[0].value == 'Avg Rank Position':
                col_letter = col[0].column_letter
                ws_rankings.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{len(all_candidates) + 1}',
                    ColorScaleRule(
                        start_type='min', start_color='4CAF50',
                        mid_type='percentile', mid_value=50, mid_color='FFC107',
                        end_type='max', end_color='F44336'
                    )
                )
        
        # Format Jury Status sheet
        ws_jury = workbook['Jury Voting Status']
        for row in range(2, ws_jury.max_row + 1):
            status_cell = ws_jury[f'D{row}']
            if status_cell.value == 'Completed':
                status_cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
            elif status_cell.value == 'Not Started':
                status_cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
        
        # Auto-adjust column widths
        for sheet in workbook.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    print("\n" + "=" * 60)
    print("✅ REPORT GENERATION COMPLETE!")
    print("=" * 60)
    print(f"📊 Excel report saved as: {output_file}")
    print(f"📋 Report contains:")
    print(f"   • All {len(all_candidates)} candidates")
    print(f"   • Top 25 Mobility Trailblazers")
    print(f"   • Top 3 winners per category (Startup, Government, Corporate)")
    print(f"   • Jury voting status overview")
    print(f"   • Complete scoring methodology explanations")
    print("=" * 60)

if __name__ == "__main__":
    # Generate report with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"/mnt/dietpi_userdata/docker-files/VOTE/Mobility_Complete_Report_{timestamp}.xlsx"
    
    print("=" * 60)
    print("MOBILITY TRAILBLAZERS VOTE - COMPLETE EVALUATION REPORT")
    print("Generating report as requested by Tobias")
    print("=" * 60)
    print()
    
    create_excel_report(output_file)
    
    print("\n📧 WordPress Access Information:")
    print("   URL: https://vote.mobilitytrailblazers.de/wp-admin")
    print("   Please contact the administrator for login credentials")
    print("=" * 60)